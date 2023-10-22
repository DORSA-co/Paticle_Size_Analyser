if __name__ == '__main__':
    import os
    import sys
    sys.path.append(os.getcwd())

import openpyxl #3.1.2
from openpyxl.cell.cell import Cell
from openpyxl.chart import  Reference, Series, BarChart
from openpyxl.worksheet.worksheet import Worksheet
import copy
#from openpyxl.chart.series import Series

from backend.Processing.Report import Report

class excelExporter:
    MAX_ROW = 100
    MAX_COL = 100
    def __init__(self,path_format:str,) -> None:
        self.file_type = '.xlsx'
        self.path_format = path_format
        self.workbook = openpyxl.load_workbook(filename = self.path_format)
        self.sheet_name:str = self.workbook.sheetnames[0]
        self.sheet:Worksheet = self.workbook[self.sheet_name]
        self.sheet_format:Worksheet = copy.deepcopy(self.sheet)
        self.charts_data = {}
        self.__find_codes__()
    
    def __find_codes__(self,):
        self.codes = {}
        for i in range(1,self.MAX_ROW):
            for j in range(1,self.MAX_COL):
                value = self.sheet.cell(row=i, column=j).value
                if value is not None:
                    value = value.strip()

                    if len(value) > 3:
                        if value[0] == '%' and value[-1] == '%':
                            self.codes[value.upper()] = (i,j)
        

    def copy_style(self,cell1:Cell, cell2:Cell):
        """copy style of cell1 into cell2
        """
        if cell1.has_style:
            cell2.font = cell1.font.copy()
            cell2.border = cell1.border.copy()
            cell2.fill = cell1.fill.copy()
            cell2.number_format = cell1.number_format
            cell2.protection = cell1.protection.copy()
            cell2.alignment = cell1.alignment.copy()

    def save(self,path):
        try:
            if len(path) > len(self.file_type):
                if path[-len(self.file_type):] != self.file_type:
                    path = path + self.file_type
            else:
                path = path + self.file_type

            self.workbook.save(path)
            return True, None
        except Exception as e:
            return False, str(e)
    

    def set_list_values(self, datas:list, start_pos:tuple[int], oriation='v') -> int:
        """set list values into excel, this function inster rows or colums from start point

        Args:
            datas (list): list of data
            start_pos (tuple[int]): _description_
            oriation (str, optional): _description_. Defaults to 'v'.

        Returns:
            tuple: end cell
        """
        i,j = start_pos

        for data in datas:
            self.copy_style(self.sheet.cell(*start_pos), 
                            self.sheet.cell(i,j))
            self.sheet.cell(i,j).value = data
            if oriation.lower() == 'h':
                j+=1
            elif oriation.lower() == 'v':
                i+=1
        
        if oriation.lower() == 'h':
            j-=1
        elif oriation.lower() == 'v':
            i-=1
        return (i,j)


    def get_chart_serie_cells(self, serie:Series) -> tuple[ tuple, tuple]:
        """returns first x cell and y cell of a chart series

        Args:
            series (Series): table serie

        Returns:
            tuple: x cell position, y cell position
        """
        x_refrence = serie.tx.strRef.f
        y_refrence = serie.val.numRef.ref

        x_cell = self.chart_refrence_to_cell_index(x_refrence)
        y_cell = self.chart_refrence_to_cell_index(y_refrence)
        return x_cell , y_cell


    def chart_refrence_to_cell_index(self, ref:str) -> tuple:
        """extract firts cell of chart data refrence

        Args:
            ref (str): _description_

        Returns:
            tuple: cell index (row, col)
        """
        #ref example for one cell 'Sheet1!$A$10' for multi cell 'Sheet1!$A$10'  
        ref = ref.replace(self.sheet_name, '')
        ref = ref.replace('$', '')
        ref = ref.replace('!', '')

        if ':' in ref:
            idx = ref.find(':')
            ref = ref[:idx]
        cell = self.sheet[ref]
        return cell.row, cell.column


    def make_chart_refrence(self, start_cell, end_cell) -> str:
        #ref example 'Sheet1!$A$11:$B$11'
        start_cell = self.sheet.cell(*start_cell)
        end_cell = self.sheet.cell(*end_cell)
        ref = f'{self.sheet_name}!'
        ref = ref + f'${start_cell.column_letter}${start_cell.row}'
        ref = ref + ':'
        ref = ref + f'${end_cell.column_letter}${end_cell.row}'
        return ref
    
    def set_chart_ref(self, serie, x_refrence, y_refrence):
        serie.tx.strRef.f = x_refrence
        serie.val.numRef.ref = y_refrence




class reportExcelExporter(excelExporter):
    MAX_COL = 50
    MAX_ROW = 100
    

    def __update_pos(self, start_pos, insert_col, insert_row):
        for code, pos in self.codes.items():
            pos = list(pos)
            if pos[0] > start_pos[0]:
                pos[0] = pos[0] + insert_row
            
            if pos[1] > start_pos[1]:
                pos[1] = pos[1] + insert_col
            
            pos = tuple(pos)
            self.codes[code] = pos
            
    def render_charts(self, datas_end_cell: dict):
        charts = self.sheet._charts
        for chart in charts:
            series = chart.series
            for serie in series:
                x_start_cell_pos, y_start_cell_pos = self.get_chart_serie_cells(serie)
                x_cell_code = self.sheet_format.cell(*x_start_cell_pos).value
                y_cell_code = self.sheet_format.cell(*y_start_cell_pos).value

                if x_cell_code in datas_end_cell.keys() and y_cell_code in datas_end_cell.keys():
                    x_end_cell_pos = datas_end_cell[x_cell_code] 
                    y_end_cell_pos = datas_end_cell[y_cell_code]

                    y_ref = self.make_chart_refrence(y_start_cell_pos, y_end_cell_pos)
                    x_ref = self.make_chart_refrence(x_start_cell_pos, x_end_cell_pos)

                    self.set_chart_ref(serie, x_ref, y_ref)





    def render(self, report:Report):
        range_statistics = report.get_ranges_statistics()
        global_statistics = report.get_global_statistics()
        
        wrong_codes = []
        self.datas_end_cell = {}
        
        
        for code,pos in self.codes.items():
            datas = None

            if code == '%NAME%':
                self.sheet.cell(*pos).value = report.name

            elif code == '%DATE%':
                self.sheet.cell(*pos).value = report.date.strftime('%Y-%M-%D')
            
            elif code == '%TIME%':
                self.sheet.cell(*pos).value = report.time.strftime('%H:%m')

            elif code == '%USER%':
                self.sheet.cell(*pos).value = report.username
            
            elif code == '%STANDARD%':
                self.sheet.cell(*pos).value = report.standard['name']
            
            elif code == '%TOTAL_AVRAGE%':
                self.sheet.cell(*pos).value = global_statistics['avrage']
            
            elif code == '%TOTAL_STD%':
                self.sheet.cell(*pos).value = global_statistics['std']
            
            elif code in ['%RANGE_NAME_VERTICALLY%', '%RANGE_NAME_HORIZONTAL%']:
                datas = report.ranges_string
            
            elif code in ['%RANGE_PERCENT_VERTICALLY%', '%RANGE_PERCENT_HORIZONTAL%']:
                datas = list(map(lambda x:x['percent'], range_statistics))
            
            elif code in ['%RANGE_AVRAGE_VERTICALLY%', '%RANGE_AVRAGE_HORIZONTAL%']:
                datas = list(map(lambda x:x['avrage'], range_statistics))

            elif code in ['%RANGE_STD_VERTICALLY%', '%RANGE_STD_HORIZONTAL%']:
                datas = list(map(lambda x:x['std'], range_statistics))

            else:
                wrong_codes.append(code)

            if datas is not None:
                end_cell = None
                if 'VERTICALLY' in code:
                    end_cell = self.set_list_values(datas, pos, oriation='v')

                elif 'HORIZONTAL' in code:
                    end_cell = self.set_list_values(datas, pos, oriation='h')
                
                self.datas_end_cell[code] = end_cell

        self.render_charts( self.datas_end_cell)

        return wrong_codes   

    



    


# if __name__ == '__main__':
#     from backend.Utils.StorageUtils import objectSaver

#     report = objectSaver.load(r'C:\Users\amir\AppData\Local\Dorsa-PSA-Reports\20231021_1524_admin_20231021_152422\report')
#     excel = reportExcelExporter(r'files\export_formats\report_format.xlsx')
#     excel.render(report)
#     excel.save('test')
    
#     pass